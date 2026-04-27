"""Extract rental + sales medians from Excel into Parquet + DuckDB.

Ported from the upstream isochrones project's
`scripts/rental_sales/extract.py` (commit bdc2961, tag v2.1.0). The
schema mapping (`etl/rental_sales_schema.yaml`) lists each source xlsx
file with its sheets, cell ranges, and per-sheet metadata (dwelling
type, bedroom count, statistic columns). For each file we walk the
configured sheets, cross-product the time-bucket columns with the
geospatial-name rows, and emit one row per (suburb, time, dwelling
type, bedrooms, statistic) cell.

Differences from upstream
- LGA-granularity entries are omitted from the YAML for now (the app
  only renders suburb boundaries today). Skip with a log message if
  encountered.
- Reads SAL parquet from the path produced by our own `etl extract sal`
  step rather than the upstream's pre-converted location.
- Outputs straight to public/data/rental_sales.duckdb (where the
  frontend reads from) plus a parquet checkpoint under data/converted/.

Special-case suburb-name remappings preserved verbatim from upstream
because the source data uses inconsistent forms (Merri-bek's hyphen,
Mornington Peninsula's apostrophe-contraction, etc.).
"""

from __future__ import annotations

import datetime as dt
import logging
from pathlib import Path
from typing import Any

import duckdb
import geopandas as gpd
import openpyxl as xl
import pandas as pd
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet
from ruamel.yaml import YAML

log = logging.getLogger("etl.steps.extract_rental_sales")

# Suburb names where the source data uses a non-standard form. Maps the
# raw value seen in the xlsx (lowercased) to the SAL_NAME21 lookup key
# (also lowercased). Ported verbatim from upstream's special cases.
SUBURB_NAME_REMAPS: dict[str, list[str]] = {
    "merri-bek": ["merri-bek"],  # the dash is genuine, not a delimiter
    "mornington penin'a": ["mornington peninsula"],
    "colac-otway": ["colac otway"],
    "east brunswick": ["brunswick east"],
    "west brunswick": ["brunswick west"],
    "east st kilda": ["st kilda east"],
    "west st kilda": ["st kilda west"],
}

# Aggregate "Victoria total / metro / non-metro" rows in the source
# spreadsheets — geospatial values that should be skipped.
SKIP_GEO_VALUES = frozenset({"Group Total", "Grand Total", "Victoria", "Metro", "Non-Metro"})


def _build_sal_lookup(sal_parquet: Path) -> dict[str, str]:
    """Map lowercased SAL_NAME21 -> SAL_CODE21. Strips '(vic.)' suffix."""
    if not sal_parquet.exists():
        raise FileNotFoundError(
            f"SAL parquet not found at {sal_parquet}. Run `etl extract sal` first."
        )
    gdf = gpd.read_parquet(sal_parquet)
    if "SAL_NAME21" not in gdf.columns or "SAL_CODE21" not in gdf.columns:
        raise ValueError(
            f"SAL parquet missing SAL_NAME21 / SAL_CODE21 columns. Available: {list(gdf.columns)}"
        )
    return {
        str(row["SAL_NAME21"]).lower().replace(" (vic.)", ""): str(row["SAL_CODE21"])
        for _, row in gdf[["SAL_NAME21", "SAL_CODE21"]].iterrows()
    }


def _split_geo_value(geo_value: str) -> list[str]:
    """Split a hyphen-grouped geo name into its constituent lookup keys.

    The source xlsx has values like "Albert Park-Middle Park-West St Kilda"
    where the dash is a delimiter. But some names contain a real dash
    ("Merri-bek", "Colac-Otway") — those are handled by the remap table.
    """
    lowered = geo_value.lower()
    if lowered in SUBURB_NAME_REMAPS:
        return SUBURB_NAME_REMAPS[lowered]
    return [v.strip() for v in lowered.split("-") if v.strip()]


def _process_sheet(
    sheet: Worksheet,
    file_config: dict[str, Any],
    sheet_config: dict[str, Any],
    sal_lookup: dict[str, str],
) -> list[dict[str, Any]]:
    statistics = list(sheet_config["statistic"])
    time_bucket_format = file_config["time_bucket_format"]
    geospatial_type = file_config["data_granularity"]
    data_type = file_config["data_type"]
    data_frequency = file_config["data_frequency"]
    source_file = file_config["file"]

    tb_start_col, tb_start_row, tb_end_col, _ = range_boundaries(sheet_config["time_bucket_range"])
    geo_start_col, geo_start_row, _, geo_end_row = range_boundaries(
        sheet_config["geospatial_range"]
    )

    bedrooms_str = str(sheet_config["bedrooms"])
    dwelling_type = sheet_config["dwelling_type"]

    rows: list[dict[str, Any]] = []
    for geo_row in range(geo_start_row, geo_end_row + 1):
        geo_value = sheet.cell(row=geo_row, column=geo_start_col).value
        if not geo_value or geo_value in SKIP_GEO_VALUES:
            continue

        geo_keys = _split_geo_value(str(geo_value))
        geo_codes = [sal_lookup[k] for k in geo_keys if k in sal_lookup]

        for time_col in range(tb_start_col, tb_end_col + 1):
            time_value = sheet.cell(row=tb_start_row, column=time_col).value
            if not time_value:
                continue
            stat_index = (time_col - tb_start_col) % len(statistics)
            stat_type = statistics[stat_index]
            cell_value = sheet.cell(row=geo_row, column=time_col).value
            if cell_value is None or cell_value in ("-", ""):
                continue
            try:
                value_float = float(cell_value)
            except (ValueError, TypeError):
                continue
            try:
                time_bucket = dt.datetime.strptime(str(time_value), time_bucket_format).date()
            except ValueError:
                continue
            rows.append(
                {
                    "geospatial": str(geo_value),
                    "geospatial_codes": "-".join(geo_codes),
                    "geospatial_type": geospatial_type,
                    "time_bucket": time_bucket,
                    "dwelling_type": dwelling_type,
                    "bedrooms": bedrooms_str,
                    "dwelling_class": f"{dwelling_type}-{bedrooms_str}",
                    "statistic": stat_type,
                    "value": value_float,
                    "data_type": data_type,
                    "data_frequency": data_frequency,
                    "source_file": source_file,
                    "source_sheet": sheet.title,
                    "cell": f"{xl.utils.get_column_letter(time_col)}{geo_row}",
                }
            )
    return rows


def _process_file(
    file_path: Path,
    file_config: dict[str, Any],
    sal_lookup: dict[str, str],
) -> list[dict[str, Any]]:
    log.info("Processing file: %s", file_path.name)
    # NOT read_only — random `.cell(row, col)` access is O(N) per lookup in
    # read-only mode (it walks the sheet stream each time). Our iteration
    # pattern is column-major over the time-bucket range, so read-only made
    # the extract O(N^2) and effectively hung. Plain mode loads the full
    # sheet into memory but each xlsx is only a few MB.
    workbook = xl.load_workbook(file_path, data_only=True)
    configured_sheets = {item["sheet"]: item for item in file_config["sheets"]}
    rows: list[dict[str, Any]] = []
    for sheet_name in workbook.sheetnames:
        sheet_config = configured_sheets.get(sheet_name)
        if not sheet_config:
            log.debug("  No config for sheet %r, skipping", sheet_name)
            continue
        log.info("  Sheet: %s", sheet_name)
        sheet_rows = _process_sheet(workbook[sheet_name], file_config, sheet_config, sal_lookup)
        log.info("    -> %d rows", len(sheet_rows))
        rows.extend(sheet_rows)
    workbook.close()
    return rows


def run(
    *,
    input_dir: Path,
    schema_file: Path,
    sal_parquet: Path,
    output_parquet: Path,
    output_duckdb: Path,
) -> int:
    """Build the rental_sales DuckDB + parquet from xlsx sources via the schema map."""
    if not input_dir.exists():
        raise FileNotFoundError(f"Rental-sales source dir not found: {input_dir}")
    if not schema_file.exists():
        raise FileNotFoundError(f"Schema mapping not found: {schema_file}")

    log.info("Loading SAL lookup from %s", sal_parquet)
    sal_lookup = _build_sal_lookup(sal_parquet)
    log.info("  -> %d SAL entries", len(sal_lookup))

    log.info("Loading schema mapping: %s", schema_file)
    config = YAML(typ="safe").load(schema_file.read_text())

    all_rows: list[dict[str, Any]] = []
    for file_config in config:
        if file_config["data_granularity"] != "suburb":
            log.info(
                "Skipping non-suburb-granularity file: %s (granularity=%r)",
                file_config["file"],
                file_config["data_granularity"],
            )
            continue
        file_path = input_dir / file_config["file"]
        if not file_path.exists():
            log.warning("Source file missing, skipping: %s", file_path)
            continue
        all_rows.extend(_process_file(file_path, file_config, sal_lookup))

    log.info("Total rows extracted: %d", len(all_rows))
    df = pd.DataFrame(all_rows)
    if df.empty:
        raise ValueError("No rows extracted — check input dir + schema mapping")

    output_parquet.parent.mkdir(parents=True, exist_ok=True)
    df.to_parquet(output_parquet, index=False)
    log.info(
        "Wrote parquet: %s (%.2f MB)", output_parquet, output_parquet.stat().st_size / 1_048_576
    )

    output_duckdb.parent.mkdir(parents=True, exist_ok=True)
    if output_duckdb.exists():
        output_duckdb.unlink()
    con = duckdb.connect(str(output_duckdb))
    try:
        con.register("source_df", df)
        con.execute("CREATE TABLE rental_sales AS SELECT * FROM source_df")
        row = con.execute("SELECT COUNT(*) FROM rental_sales").fetchone()
        n = row[0] if row else 0
        log.info("Wrote DuckDB: %s (%d rows)", output_duckdb, n)
    finally:
        con.close()

    return len(df)
